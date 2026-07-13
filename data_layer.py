from dataclasses import dataclass, field
from datetime import datetime,date
from typing import Literal, Optional, Callable
from uuid import uuid4
import openpyxl


@dataclass
class Expense:
    expense_id: str
    category: str
    description: str
    amount: float
    expense_date: date
    vendor: str



@dataclass
class Invoice:
    invoice_id: str
    client_name: str
    amount: float
    issue_date: date
    due_date: date
    status: str = "Pending"
    days_overdue: int = 0


@dataclass
class Payment:
    payment_id: str
    invoice_id: str
    amount_paid: float
    payment_date: date
    payment_method: str 


def load_invoices():
    wb = openpyxl.load_workbook("accounts.xlsx", data_only=True)
    ws = wb["Invoices"]
    
    invoices = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        invoice_id, client_name, amount, issue_date, due_date, status, days_overdue = row
        invoice = Invoice(invoice_id=invoice_id, client_name=client_name, amount=amount, issue_date=_to_date(issue_date), due_date=_to_date(due_date), status=status, days_overdue=days_overdue)
        invoices.append(invoice)
    wb.close()
    
    return invoices
def _to_date(value):
    if isinstance(value, datetime):
        return value.date()
    return value
def load_expenses():
    wb = openpyxl.load_workbook("accounts.xlsx", data_only=True)
    ws = wb["Expenses"]
    
    expenses = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        expense_id, category, description, amount, expense_date, vendor = row
        expense = Expense(expense_id=expense_id, category=category, description=description, amount=amount, expense_date=expense_date, vendor=vendor)
        expenses.append(expense)
    wb.close()
    
    return expenses

def load_payments():
    
    wb = openpyxl.load_workbook("accounts.xlsx", data_only=True)
    ws = wb["Payments"]
    
    payments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        payment_id, invoice_id, amount_paid, payment_date, payment_method = row
        payment = Payment(payment_id=payment_id, invoice_id=invoice_id, amount_paid=amount_paid, payment_date=payment_date, payment_method=payment_method)
        payments.append(payment)
    wb.close()
    
    return payments
